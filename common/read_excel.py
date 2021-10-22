# !/user/bin/env python
# _*_ coding: utf-8 _*_
# @Time     : 2021/5/20 13:48
# @Author   : 王俊
# @File     : read_excel.py
# @Software : PyCharm
import os
import openpyxl
from common import DATA_DIR
from numpy.random._common import namedtuple


class ExcelData(object):
    pass


class ReadExcelData(object):
    __instance = None

    def __new__(cls, *args, **kwargs):
        if cls.__instance is None:
            cls.__instance = super().__new__(cls)
        return cls.__instance

    def __init__(self, f, sheetname=None):
        self.f = os.path.join(DATA_DIR, f)
        self.sheetname = sheetname
        self.open()

    def open(self):
        self.workbook = openpyxl.load_workbook(self.f)
        self.sheet = self.workbook.active if self.sheetname is None else self.workbook[self.sheetname]

    def save_excel(self):
        self.workbook.save(self.f)
        self.workbook.close()

    def get_data(self) -> list:
        """
            cases = []
        for row in rows[1:]:
            data = []
            # 循环遍历第一行之外的数据，聚合打包成字典
            for r in row:
                # 将格子中的数据添加到data中
                data.append(r.value)
            # 使用zip函数打包成字典
            case = dict(zip(title, data))
            # 将每条用例添加到整个用例列表
            cases.append(case)
            # 关闭工作簿
            self.close()
            # 返回用例
        return cases
        :return:[{}]
        """
        rows = list(self.sheet.rows)
        cases = [dict(zip([row.value for row in rows[0]], [r.value for r in row])) for row in rows[1:]]
        return cases

    def get_data_obj(self) -> list:
        """
            采用对象-->属性-->属性值方式进行用例存储
        :return:
        """
        rows = list(self.sheet.rows)
        cases = []
        for row in rows[1:]:
            case_obj = ExcelData()
            for k, v in list(zip([row.value for row in rows[0]], [r.value for r in row])):
                setattr(case_obj, k, v)
            cases.append(case_obj)
        return cases

    def get_name_tuple_data(self, row=None) -> list:
        self.sheet_head_tuple = tuple(self.sheet.iter_rows(min_row=1, values_only=True))[0]
        self.data_tuple = namedtuple("data_tuple", self.sheet_head_tuple)
        if row is None:
            cases = [self.data_tuple(*_data) for _data in self.sheet.iter_rows(min_row=2, values_only=True)]
            return cases
        else:
            return self.get_specific_case_data(row=row)

    def get_specific_case_data(self, row) -> list:
        """
            获取指定行的测试用例
        :param row: 行号
        :return:
        """
        if isinstance(row, int) and (1 <= row <= self.sheet.max_row):
            return self.data_tuple(*tuple(self.sheet.iter_rows(min_row=row + 1, max_row=row + 1, values_only=True))[0])
        else:
            raise TypeError("行号只能为整数并且大于1！")

    def write_data(self, row, column, value):
        """
            将数据写入excel中
        :param row: 行
        :param column:列
        :param value: 需要写入的值
        :return:
        """
        self.sheet.cell(row=row, column=column, value=value)
        self.save_excel()

    def close(self):
        self.workbook.close()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_tb is not None or exc_val is not None:
            print(exc_tb)
            print(exc_val)
        return self.close()


if __name__ == '__main__':
    with ReadExcelData("order_cases.xlsx", "createorder") as r:
        # r.get_data()
        # r.get_data_obj()
        print(r.get_data())
        # print(r.get_data_obj()[0].data)
        # print(r.get_name_tuple_data())
